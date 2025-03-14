import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions

driver = webdriver.Chrome()
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.implicitly_wait(10)   # 5 seconds is max time out

def updateExcelData(filePath, searchTerm, colName, newValue):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    dict = {}

    for i in range(1, sheet.max_column+1):
        if sheet.cell(row = 1, column = i).value == colName:
            dict["col"] = i

    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row = i, column = j).value == searchTerm:
                dict["row"] = i

    # Edit the excel with updated value
    sheet.cell(row=dict["row"], column=dict["col"]).value = newValue
    book.save(filePath)  # save modification


filePath = "C:/Users/user/Downloads/download.xlsx"
fruitName = "Apple"
newValue = "999"

driver.find_element(By.ID, "downloadButton").click()

updateExcelData(filePath, "Apple", "price", newValue)

#upload
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(filePath)

wait = WebDriverWait(driver, 10)
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
actualPrice = driver.find_element(By.XPATH, "//div[text()='"+fruitName+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text

print(actualPrice)

assert actualPrice == newValue